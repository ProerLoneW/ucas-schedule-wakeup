#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ucas_crawler.py — 国科大课表爬取 + WakeUp 课程表导出（skill 配套工具）
====================================================================

子命令：
  doctor                              检查依赖（playwright/openpyxl/Chromium）
  login-browser [--relogin] [--timeout N]
                                      弹出浏览器让用户登录学习通（扫码/密码均可），
                                      自动检测登录成功并保存 Cookie（含 HttpOnly）
  save-cookie --cookie "整串Cookie"   保存用户从浏览器 F12 复制的 Cookie
  check                               检查当前登录态是否有效
  crawl [--out DIR] [--no-merge] [--cookie "STR"]
                                      爬取全部学期（逐周扫描+去重），每学期导出
                                      WakeUp 格式的 .xlsx 和 .csv，另存原始 JSON
  export --json schedule_all.json [--out DIR]
                                      用已保存的原始 JSON 重新生成 WakeUp 文件

登录态文件保存在 ~/.ucas_schedule/（cookies.json / cookie.txt），跨目录可复用。

技术要点（来自对该站前端 JS 的逆向分析）：
  · 课表接口 GET /pc/curriculum/getMyLessons 需要登录态（含 HttpOnly Cookie，
    如 vc3；仅靠 document.cookie 复制的 Cookie 会被 302 踢回登录页）
  · 不带 week 参数只返回"当前周"课程；必须 week=1..maxWeek 逐周扫描，
    再按 lessonConfigUuid 去重合并得到整学期课程
  · WakeUp 导入格式为固定 7 列：课程名称|星期|开始节数|结束节数|老师|地点|周数，
    周数支持 1-16、7-11单、2-16双、多段用中文顿号连接（官方文档
    https://wakeup.fun/doc/import_from_csv.html）
"""

import argparse
import csv
import importlib.util
import json
import ssl
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

KB = "https://kb.mooc.ucas.edu.cn"
PASSPORT = "https://passport.mooc.ucas.edu.cn"
SCHEDULE_URL = f"{KB}/res/pc/curriculum/schedule.html"
LOGIN_URL = f"{PASSPORT}/login?fid=&newversion=true&refer={SCHEDULE_URL}"
STATE_DIR = Path.home() / ".ucas_schedule"
COOKIES_JSON = STATE_DIR / "cookies.json"
COOKIE_TXT = STATE_DIR / "cookie.txt"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36")

WAKEUP_HEADERS = ["课程名称", "星期", "开始节数", "结束节数", "老师", "地点", "周数"]


# ================================================================ 依赖检查
def have_module(name: str) -> bool:
    return importlib.util.find_spec(name) is not None


def cmd_doctor(args=None):
    problems = []
    if not have_module("openpyxl"):
        problems.append(("openpyxl", "pip3 install openpyxl"))
    if not have_module("playwright"):
        problems.append(("playwright(仅浏览器登录需要)", "pip3 install playwright"))
    else:
        r = subprocess.run([sys.executable, "-m", "playwright", "install",
                            "chromium", "--dry-run"],
                           capture_output=True, text=True)
        if r.returncode != 0:
            problems.append(("Chromium内核(仅浏览器登录需要)",
                             "python3 -m playwright install chromium"))
    if problems:
        print("[缺少依赖]")
        for name, fix in problems:
            print(f"  · {name}: {fix}")
        sys.exit(1)
    print("[OK] 依赖齐全（openpyxl、playwright、Chromium）")


# ================================================================ 登录态管理
def load_cookie_header(cli_cookie: str = "") -> str:
    """优先级：命令行传入 > cookie.txt（F12 复制） > cookies.json（浏览器登录导出）。"""
    if cli_cookie:
        return cli_cookie.strip()
    if COOKIE_TXT.exists() and COOKIE_TXT.read_text(encoding="utf-8").strip():
        return COOKIE_TXT.read_text(encoding="utf-8").strip()
    if COOKIES_JSON.exists():
        try:
            cookies = json.loads(COOKIES_JSON.read_text(encoding="utf-8"))
            header = "; ".join(f"{c['name']}={c['value']}" for c in cookies)
            if header:
                return header
        except Exception:
            pass
    return ""


def cmd_save_cookie(args):
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    COOKIE_TXT.write_text(args.cookie.strip(), encoding="utf-8")
    print(f"[OK] Cookie 已保存 -> {COOKIE_TXT}")
    print("     正在验证登录态……")
    return cmd_check(args)


def cmd_login_browser(args):
    if not have_module("playwright"):
        sys.exit("[错误] 未安装 playwright。pip3 install playwright && "
                 "python3 -m playwright install chromium，或改用 Cookie 方式。")
    from playwright.sync_api import sync_playwright
    profile = STATE_DIR / "browser_profile"
    with sync_playwright() as p:
        ctx = p.chromium.launch_persistent_context(
            str(profile), headless=False, user_agent=UA,
            viewport={"width": 1280, "height": 800},
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = ctx.pages[0] if ctx.pages else ctx.new_page()
        if not args.relogin:
            # 尝试复用上次登录态
            if COOKIES_JSON.exists():
                try:
                    ctx.add_cookies(json.loads(COOKIES_JSON.read_text(encoding="utf-8")))
                except Exception:
                    pass
            if _browser_logged_in(ctx):
                print("[OK] 浏览器档案里的登录态仍有效，无需重新登录")
                _persist_cookies(ctx)
                ctx.close()
                return
        page.goto(LOGIN_URL)
        print("=" * 56)
        print("浏览器已打开登录页，请在浏览器中完成登录（任选其一）：")
        print("  · 学习通 App 扫码（推荐，不易触发验证码）")
        print("  · 手机号/超星号 + 学习通密码")
        print("  · 其他登录方式 -> 机构账号（中国科学院大学 + 学号 + 学习通密码）")
        print("  · 验证码登录（手机短信）")
        print(f"等待登录中……（超时 {args.timeout} 秒，成功后自动继续）")
        print("=" * 56)
        deadline = time.time() + args.timeout
        while time.time() < deadline:
            if _browser_logged_in(ctx):
                print("[OK] 检测到登录成功！")
                try:
                    page.goto(SCHEDULE_URL, wait_until="domcontentloaded")
                except Exception:
                    pass
                _persist_cookies(ctx)
                ctx.close()
                return
            time.sleep(2)
        ctx.close()
        sys.exit("[超时] 未检测到登录。重新运行，或加大 --timeout。")


def _browser_logged_in(ctx) -> bool:
    resp = ctx.request.get(f"{KB}/pc/curriculum/getMyLessons",
                           params={"curTime": int(time.time() * 1000)},
                           headers={"User-Agent": UA, "Referer": SCHEDULE_URL})
    try:
        return json.loads(resp.text()).get("result") == 1
    except Exception:
        return False


def _persist_cookies(ctx):
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    cookies = ctx.cookies()
    COOKIES_JSON.write_text(json.dumps(cookies, ensure_ascii=False, indent=2),
                            encoding="utf-8")
    header = "; ".join(f"{c['name']}={c['value']}" for c in cookies)
    COOKIE_TXT.write_text(header, encoding="utf-8")
    print(f"[OK] 已导出 {len(cookies)} 条 Cookie -> {STATE_DIR}")
    print("     下次运行自动复用，无需重新登录（失效时会提示）")


def cmd_check(args):
    header = load_cookie_header(getattr(args, "cookie", ""))
    if not header:
        print("[未登录] 没有可用的登录态。请先选择认证方式：")
        print("  A. save-cookie --cookie \"...\"   （浏览器 F12 复制 Cookie）")
        print("  B. login-browser                 （弹浏览器登录学习通）")
        sys.exit(2)
    body = http_get_json("/pc/curriculum/getMyLessons", {}, header)
    if body.get("result") == 1:
        curr = body["data"].get("curriculum", {})
        print(f"[OK] 登录态有效：{curr.get('userName')}，"
              f"当前学期 {curr.get('schoolYear')} 第 {curr.get('semester')} 学期")
        return True
    print(f"[无效] 登录态不可用：{_auth_fail_reason(body)}")
    print("常见原因：Cookie 复制不完整（缺 HttpOnly 项，必须从 Network 请求头复制）、"
          "或登录已过期。请重新获取，或换另一种认证方式。")
    sys.exit(2)


# ================================================================ HTTP 抓取
def _ssl_context():
    """macOS 自带 Python 常缺根证书：优先 certifi，其次系统默认，最后降级不校验。"""
    try:
        import certifi
        return ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        pass
    return ssl.create_default_context()


_SSL_FALLBACK = None  # 首次证书校验失败后置 True，避免每个请求重复失败


def http_get_json(path, params, cookie_header):
    global _SSL_FALLBACK
    qs = params.copy()
    qs.setdefault("curTime", int(time.time() * 1000))
    url = KB + path + "?" + urllib.parse.urlencode(qs)
    headers = {"User-Agent": UA, "Referer": SCHEDULE_URL,
               "X-Requested-With": "XMLHttpRequest", "Cookie": cookie_header}
    for attempt in range(2):
        try:
            req = urllib.request.Request(url, headers=headers)
            if _SSL_FALLBACK:
                ctx = ssl._create_unverified_context()
                with urllib.request.urlopen(req, timeout=20, context=ctx) as resp:
                    text = resp.read().decode("utf-8", "replace")
            else:
                with urllib.request.urlopen(req, timeout=20) as resp:
                    text = resp.read().decode("utf-8", "replace")
            break
        except urllib.error.HTTPError as e:
            return {"result": 0, "error": f"HTTP {e.code}"}
        except Exception as e:
            if attempt == 0 and "CERTIFICATE_VERIFY_FAILED" in str(e):
                print("[警告] 系统 Python 缺少根证书，已降级为跳过证书校验"
                      "（可运行 /Applications/Python*/Install\\ Certificates.command 修复）")
                _SSL_FALLBACK = True
                continue
            return {"result": 0, "error": str(e)}
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # 未登录时服务端 302 到登录页，拿到的是 HTML
        return {"result": 0, "notLoggedIn": True}


def _auth_fail_reason(body) -> str:
    if body.get("error"):
        return f"网络/服务器错误：{body['error']}"
    if body.get("notLoggedIn"):
        return "请求被重定向到登录页（Cookie 缺 HttpOnly 凭据或已过期）"
    return f"接口返回：{body.get('data') or body.get('msg') or body}"


def crawl_all_terms(cookie_header, no_merge=False):
    """所有学期 × 逐周扫描 × 按 lessonConfigUuid 去重。"""
    hist = http_get_json("/pc/curriculum/getCurriculumHistory", {}, cookie_header)
    terms = ((hist.get("data") or {}).get("list")) or []
    if not terms:
        sys.exit(f"[错误] 学期列表获取失败（登录态可能已失效）：{hist}")
    results = []
    for t in terms:
        year, sem = t["schoolYear"], t["semester"]
        base = {"schoolYear": year, "semester": sem,
                "userSelectedTime": int(time.time() * 1000)}
        if no_merge:
            base["noMergeLesson"] = 1
        first = http_get_json("/pc/curriculum/getMyLessons", base, cookie_header)
        if first.get("result") != 1:
            print(f"[跳过] {year} 第{sem}学期：{first.get('data') or first.get('msg')}")
            continue
        data = first["data"]
        max_week = data.get("curriculum", {}).get("maxWeek") or 20
        merged, week_counts = {}, {}
        for week in range(1, max_week + 1):
            body = http_get_json("/pc/curriculum/getMyLessons",
                                 {**base, "week": week}, cookie_header)
            if body.get("result") != 1:
                week_counts[week] = -1
                continue
            lessons = body["data"].get("lessonArray") or []
            week_counts[week] = len(lessons)
            for les in lessons:
                if not isinstance(les, dict):
                    continue
                key = les.get("lessonConfigUuid") or "|".join(str(x) for x in (
                    les.get("name"), les.get("dayOfWeek"), les.get("beginNumber"),
                    les.get("length"), les.get("location")))
                merged.setdefault(key, les)
        data["lessonArray"] = list(merged.values())
        data["weekLessonCounts"] = week_counts
        results.append((year, sem, data))
        print(f"[完成] {year}-{int(year) + 1} 第{sem}学期：{len(merged)} 门课程"
              f"（扫描 {max_week} 周）")
    return results


# ================================================================ WakeUp 转换
def parse_weeks(lesson) -> set:
    """周次字段 -> 周号集合。原始 weeks 为逗号分隔字符串，如 '1,2,3,5,7,9,11'。"""
    weeks = lesson.get("weeks") or lesson.get("jc") or ""
    if isinstance(weeks, list):
        return {int(w) for w in weeks if str(w).strip().isdigit()}
    out = set()
    for part in str(weeks).split(","):
        part = part.strip()
        if part.isdigit():
            out.add(int(part))
    return out


def weeks_to_wakeup(week_set) -> str:
    """周号集合 -> WakeUp 周数格式。
    连续段 1-5；同奇偶的间隔段（1,3,5..）合并为 1-11单 / 2-10双；多段用顿号连接。"""
    if not week_set:
        return ""
    nums = sorted(week_set)
    runs = []
    start = prev = nums[0]
    step = 0  # 0 = 游程内只有 1 个数，步长未知
    for n in nums[1:]:
        if step == 0:
            step = n - prev
        if step in (1, 2) and n - prev == step:
            prev = n
            continue
        runs.append((start, prev, step if step in (1, 2) else 1))
        start = prev = n
        step = 0
    runs.append((start, prev, step if step in (1, 2) else 1))

    parts = []
    for lo, hi, st in runs:
        if hi == lo:
            parts.append(str(lo))
        elif st == 2 and lo % 2 == 1:
            parts.append(f"{lo}-{hi}单")
        elif st == 2 and lo % 2 == 0:
            parts.append(f"{lo}-{hi}双")
        else:
            parts.append(f"{lo}-{hi}")
    return "、".join(parts)


def lesson_to_rows(lesson, curriculum) -> list:
    """单条课程 -> WakeUp 行（7 列）。返回空列表表示数据不完整，跳过。"""
    name = (lesson.get("name") or "").strip()
    day = lesson.get("dayOfWeek")
    begin = lesson.get("beginNumber")
    length = lesson.get("length") or 1
    if not name or day is None or begin is None:
        return []
    # 早课节次可能为负数（晨N），换算成 WakeUp 的正数节次
    early = curriculum.get("earlyMorningSection") or 0
    if begin < 1:
        begin = begin + early if early else 1
    end = begin + length - 1
    teacher = (lesson.get("teacherName") or "").strip() or "无"
    loc = (lesson.get("onlineLocation") or lesson.get("location") or "").strip() or "无"
    weeks = weeks_to_wakeup(parse_weeks(lesson))
    if not weeks:
        return []
    return [[name, day, begin, end, teacher, loc, weeks]]


# ================================================================ 文件导出
def export_term(year, sem, data, out_dir: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    curriculum = data.get("curriculum", {})
    rows, skipped = [], 0
    for les in data.get("lessonArray") or []:
        if not isinstance(les, dict):
            continue
        r = lesson_to_rows(les, curriculum)
        if r:
            rows.extend(r)
        else:
            skipped += 1

    stem = f"WakeUp课表_{year}-{int(year) + 1}第{sem}学期"
    xlsx_path, csv_path = out_dir / f"{stem}.xlsx", out_dir / f"{stem}.csv"

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{int(year) + 1}-{sem}"
    ws.append(WAKEUP_HEADERS)
    fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for row in rows:
        ws.append(row)
    for i, w in enumerate([28, 8, 10, 10, 16, 24, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(xlsx_path)

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(WAKEUP_HEADERS)
        writer.writerows(rows)

    note = f"（{skipped} 条信息不全被跳过）" if skipped else ""
    print(f"[导出] {xlsx_path.name} / {csv_path.name}：{len(rows)} 行 {note}")
    return rows


def cmd_crawl(args):
    header = load_cookie_header(args.cookie)
    if not header:
        print("[未登录] 请先完成认证（save-cookie 或 login-browser），"
              "或用 --cookie 直接传入。")
        sys.exit(2)
    body = http_get_json("/pc/curriculum/getMyLessons", {}, header)
    if body.get("result") != 1:
        print(f"[登录态无效] {_auth_fail_reason(body)}")
        print("请重新认证（可换另一种方式），或用 --cookie 直接传入。")
        sys.exit(2)

    out_dir = Path(args.out).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    terms = crawl_all_terms(header, args.no_merge)

    raw_path = out_dir / "schedule_all.json"
    raw_path.write_text(json.dumps({f"{y}_{s}": d for y, s, d in terms},
                                   ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[导出] 原始数据 -> {raw_path}")

    exported = 0
    for year, sem, data in terms:
        if data.get("lessonArray"):
            exported += 1
            export_term(year, sem, data, out_dir)
        else:
            print(f"[跳过] {year}-{int(year) + 1} 第{sem}学期：全学期无课程，不生成文件")
    print(f"\n全部完成：{len(terms)} 个学期，{exported} 个学期有课。")
    print("导入 WakeUp：App 内 课表 -> 导入 -> Excel导入，选择 .csv 文件"
          "（.xlsx 亦可），每个学期导一次并在 App 里切换课表。")


def cmd_export(args):
    data = json.loads(Path(args.json).expanduser().read_text(encoding="utf-8"))
    out_dir = Path(args.out).expanduser().resolve() if args.out else Path(args.json).parent
    out_dir.mkdir(parents=True, exist_ok=True)
    n = 0
    for key, term_data in data.items():
        if not term_data.get("lessonArray"):
            print(f"[跳过] {key}：无课程")
            continue
        year, sem = key.rsplit("_", 1)
        export_term(year, sem, term_data, out_dir)
        n += 1
    print(f"完成：{n} 个学期")


# ================================================================ 入口
def main():
    ap = argparse.ArgumentParser(description="国科大课表爬取 + WakeUp 导出")
    sub = ap.add_subparsers(dest="cmd", required=True)

    sub.add_parser("doctor", help="检查依赖")

    p = sub.add_parser("login-browser", help="弹浏览器登录学习通")
    p.add_argument("--relogin", action="store_true", help="强制重新登录")
    p.add_argument("--timeout", type=int, default=300, help="等待登录秒数")

    p = sub.add_parser("save-cookie", help="保存 F12 复制的 Cookie 并验证")
    p.add_argument("--cookie", required=True, help="整串 Cookie")

    sub.add_parser("check", help="检查登录态")

    p = sub.add_parser("crawl", help="爬取全部学期并导出 WakeUp 文件")
    p.add_argument("--out", default=".", help="输出目录（默认当前目录）")
    p.add_argument("--no-merge", action="store_true", help="教务课程不合并")
    p.add_argument("--cookie", default="", help="直接传入 Cookie（可不保存文件）")

    p = sub.add_parser("export", help="用已保存的原始 JSON 重新生成 WakeUp 文件")
    p.add_argument("--json", required=True, help="schedule_all.json 路径")
    p.add_argument("--out", default="", help="输出目录（默认同 JSON 文件）")

    args = ap.parse_args()
    {"doctor": cmd_doctor, "login-browser": cmd_login_browser,
     "save-cookie": cmd_save_cookie, "check": cmd_check,
     "crawl": cmd_crawl, "export": cmd_export}[args.cmd](args)


if __name__ == "__main__":
    main()
